# src/app/api/endpoints/documents.py
from fastapi import (
    APIRouter,
    Depends,
    BackgroundTasks,
    UploadFile,
    File,
    HTTPException,
    Query as QueryParam,
)
from fastapi.responses import FileResponse, StreamingResponse
from sqlalchemy.orm import Session, Query, joinedload
from sqlalchemy.orm.exc import StaleDataError
from sqlalchemy import or_, func, text
from typing import List, Dict, Any, Optional
import os
import io
import csv
import glob
import zipfile
from datetime import datetime
from pydantic import BaseModel, Field, conint

from app.api.dependencies import get_db, get_current_user
from app.db import models, schemas
from app.core import background_tasks as tasks_service
from app.modules.matching import engine as matching_engine
from app.utils.auditing import log_audit_event
from app.config import settings
from app.services import document_service
from app.modules.ingestion import service as ingestion_service

router = APIRouter()

PDF_STORAGE_PATH = settings.pdf_storage_path
GENERATED_PDF_STORAGE_PATH = settings.generated_pdf_storage_path


class FailedIngestionResponse(BaseModel):
    id: int
    job_id: int
    filename: str
    document_type: str
    error_message: str
    created_at: datetime

    class Config:
        from_attributes = True


@router.get("/file/{filename}")
def get_document_file(filename: str):
    filename = os.path.basename(filename)
    
    # Check contract note folder first
    contract_note_path = os.path.join("sample_data", "contract note")
    contract_filepath = os.path.join(contract_note_path, filename)
    if os.path.exists(contract_filepath) and os.path.isfile(contract_filepath):
        return FileResponse(contract_filepath)
    
    # Check regular PDF storage
    filepath = os.path.join(PDF_STORAGE_PATH, filename)
    if os.path.exists(filepath) and os.path.isfile(filepath):
        return FileResponse(filepath)
    
    # Check generated PDF storage
    filepath = os.path.join(GENERATED_PDF_STORAGE_PATH, filename)
    if os.path.exists(filepath) and os.path.isfile(filepath):
        return FileResponse(filepath)
    
    raise HTTPException(status_code=404, detail="File not found")


@router.post("/upload", response_model=schemas.Job, status_code=202)
async def upload_documents(
    background_tasks: BackgroundTasks,
    files: List[UploadFile] = File(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    if not files:
        raise HTTPException(status_code=400, detail="No files were uploaded.")
    try:
        os.makedirs(PDF_STORAGE_PATH, exist_ok=True)
    except OSError as e:
        raise HTTPException(
            status_code=500,
            detail=f"Cannot create upload directory: {e}. Check volume mounts and permissions.",
        )
    file_data_list: List[Dict[str, Any]] = []
    for file in files:
        try:
            content = await file.read()
            safe_filename = os.path.basename(file.filename)
            file_path = os.path.join(PDF_STORAGE_PATH, safe_filename)
            with open(file_path, "wb") as buffer:
                buffer.write(content)
            file_data_list.append({"filename": safe_filename, "content": content})
        except (OSError, IOError) as e:
            raise HTTPException(
                status_code=500,
                detail=f"Failed to save file {file.filename}: {e}. Check disk space and permissions.",
            )
    job = models.Job(total_files=len(files))
    db.add(job)
    db.commit()
    db.refresh(job)
    background_tasks.add_task(
        tasks_service.process_uploaded_documents, job.id, file_data_list
    )
    return job


@router.post("/sync-sample-data", response_model=schemas.Job, status_code=202)
async def sync_sample_data(
    background_tasks: BackgroundTasks, db: Session = Depends(get_db)
):
    # Process all data in sample_data folder including contract notes, customer data, and other documents
    sample_data_path = "sample_data"
    file_data_list: List[Dict[str, Any]] = []
    
    try:
        # 1. Process contract notes (PDFs in contract note folder)
        contract_note_path = os.path.join(sample_data_path, "contract note")
        if os.path.exists(contract_note_path):
            contract_files = glob.glob(os.path.join(contract_note_path, "*.pdf"))
            for file_path in contract_files:
                try:
                    with open(file_path, "rb") as f:
                        file_data_list.append({
                            "filename": os.path.basename(file_path),
                            "content": f.read(),
                            "file_type": "contract_note",
                            "source_folder": "contract note"
                        })
                except (OSError, IOError) as e:
                    print(f"Warning: Could not read contract file {file_path}: {e}")
                    continue
        
        # 2. Process customer data (Excel files in customer_data folder)
        customer_data_path = os.path.join(sample_data_path, "customer_data")
        if os.path.exists(customer_data_path):
            excel_files = glob.glob(os.path.join(customer_data_path, "*.xlsx")) + \
                         glob.glob(os.path.join(customer_data_path, "*.xls"))
            for file_path in excel_files:
                try:
                    with open(file_path, "rb") as f:
                        file_data_list.append({
                            "filename": os.path.basename(file_path),
                            "content": f.read(),
                            "file_type": "excel_customer_data",
                            "source_folder": "customer_data"
                        })
                except (OSError, IOError) as e:
                    print(f"Warning: Could not read Excel file {file_path}: {e}")
                    continue
        
        # 3. Process loan documents (PDFs in loan_document folder)
        loan_document_path = os.path.join(sample_data_path, "loan_document")
        if os.path.exists(loan_document_path):
            loan_files = glob.glob(os.path.join(loan_document_path, "*.pdf"))
            for file_path in loan_files:
                try:
                    with open(file_path, "rb") as f:
                        file_data_list.append({
                            "filename": os.path.basename(file_path),
                            "content": f.read(),
                            "file_type": "loan_document",
                            "source_folder": "loan_document"
                        })
                except (OSError, IOError) as e:
                    print(f"Warning: Could not read loan document {file_path}: {e}")
                    continue
        
        # 4. Process loan policy documents (PDFs in loan policy folder)
        loan_policy_path = os.path.join(sample_data_path, "loan policy")
        if os.path.exists(loan_policy_path):
            policy_files = glob.glob(os.path.join(loan_policy_path, "*.pdf"))
            print(f"🔍 Found {len(policy_files)} loan policy files in {loan_policy_path}")
            for file_path in policy_files:
                try:
                    file_stat = os.stat(file_path)
                    print(f"📄 Reading loan policy: {os.path.basename(file_path)}")
                    print(f"   File path: {file_path}")
                    print(f"   File size: {file_stat.st_size} bytes")
                    print(f"   Last modified: {datetime.fromtimestamp(file_stat.st_mtime)}")
                    
                    with open(file_path, "rb") as f:
                        content = f.read()
                        print(f"   Content read: {len(content)} bytes")
                        file_data_list.append({
                            "filename": os.path.basename(file_path),
                            "content": content,
                            "file_type": "loan_policy",
                            "source_folder": "loan policy"
                        })
                except (OSError, IOError) as e:
                    print(f"Warning: Could not read loan policy {file_path}: {e}")
                    continue
                    
    except OSError as e:
        raise HTTPException(
            status_code=500, detail=f"Cannot access sample data directory: {e}"
        )
    
    if not file_data_list:
        raise HTTPException(
            status_code=404, detail="No files found in sample_data directories."
        )
    
    job = models.Job(total_files=len(file_data_list))
    db.add(job)
    db.commit()
    db.refresh(job)
    
    # Use enhanced processing function that handles multiple file types
    background_tasks.add_task(
        tasks_service.process_uploaded_documents, job.id, file_data_list
    )
    return job


@router.get("/jobs/{job_id}", response_model=schemas.Job)
def get_job_status(job_id: int, db: Session = Depends(get_db)):
    job = db.query(models.Job).filter(models.Job.id == job_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job ID not found")
    return job


@router.get("/jobs", response_model=List[schemas.Job])
def get_all_jobs(
    limit: int = 20,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    return (
        db.query(models.Job).order_by(models.Job.created_at.desc()).limit(limit).all()
    )


@router.post("/search", response_model=List[schemas.Invoice])
def search_invoices_flexible(
    request: schemas.SearchRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    return document_service.search_invoices_logic(db, request, current_user)


@router.post("/export")
def export_documents(
    request: schemas.SearchRequest,
    export_format: str = QueryParam(
        "csv", description="Export format", regex="^(csv|xlsx|pdf)$"
    ),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    invoices = document_service.search_invoices_logic(db, request, current_user)
    if not invoices:
        raise HTTPException(
            status_code=404, detail="No invoices found for the selected criteria."
        )
    all_pos, all_grns = {}, {}
    for inv in invoices:
        for po in inv.purchase_orders:
            all_pos[po.id] = po
        for grn in inv.grns:
            all_grns[grn.id] = grn
    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    if export_format == "csv":
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_f:
            inv_output = io.StringIO()
            inv_writer = csv.writer(inv_output)
            inv_header = [
                "invoice_id",
                "vendor_name",
                "invoice_date",
                "due_date",
                "status",
                "subtotal",
                "tax",
                "grand_total",
                "related_po_numbers",
                "gl_code",
            ]
            inv_writer.writerow(inv_header)
            for inv in invoices:
                inv_writer.writerow(
                    [
                        inv.invoice_id,
                        inv.vendor_name,
                        inv.invoice_date,
                        inv.due_date,
                        inv.status.value,
                        inv.subtotal,
                        inv.tax,
                        inv.grand_total,
                        ", ".join(inv.related_po_numbers or []),
                        getattr(inv, "gl_code", ""),
                    ]
                )
            zip_f.writestr("invoices.csv", inv_output.getvalue())
            if all_pos:
                po_output = io.StringIO()
                po_writer = csv.writer(po_output)
                po_header = [
                    "po_number",
                    "vendor_name",
                    "order_date",
                    "grand_total",
                    "status",
                ]
                po_writer.writerow(po_header)
                for po in all_pos.values():
                    po_writer.writerow(
                        [
                            po.po_number,
                            po.vendor_name,
                            po.order_date,
                            po.grand_total,
                            getattr(po, "status", "unknown"),
                        ]
                    )
                zip_f.writestr("purchase_orders.csv", po_output.getvalue())
            if all_grns:
                grn_output = io.StringIO()
                grn_writer = csv.writer(grn_output)
                grn_header = ["grn_number", "po_number", "received_date", "status"]
                grn_writer.writerow(grn_header)
                for grn in all_grns.values():
                    grn_writer.writerow(
                        [
                            grn.grn_number,
                            grn.po_number,
                            grn.received_date,
                            getattr(grn, "status", "unknown"),
                        ]
                    )
                zip_f.writestr("goods_receipt_notes.csv", grn_output.getvalue())
        zip_buffer.seek(0)
        return StreamingResponse(
            zip_buffer,
            media_type="application/zip",
            headers={
                "Content-Disposition": f"attachment; filename=ap_export_csv_{timestamp}.zip"
            },
        )
    elif export_format == "xlsx":
        try:
            from openpyxl import Workbook
            from openpyxl.writer.excel import save_virtual_workbook

            wb = Workbook()
            ws_inv = wb.active
            ws_inv.title = "Invoices"
            inv_headers = [
                "Invoice ID",
                "Vendor Name",
                "Invoice Date",
                "Due Date",
                "Status",
                "Subtotal",
                "Tax",
                "Grand Total",
                "Related PO Numbers",
                "GL Code",
            ]
            ws_inv.append(inv_headers)
            for inv in invoices:
                ws_inv.append(
                    [
                        inv.invoice_id,
                        inv.vendor_name,
                        inv.invoice_date,
                        inv.due_date,
                        inv.status.value,
                        inv.subtotal,
                        inv.tax,
                        inv.grand_total,
                        ", ".join(inv.related_po_numbers or []),
                        getattr(inv, "gl_code", ""),
                    ]
                )
            if all_pos:
                ws_po = wb.create_sheet("Purchase Orders")
                po_headers = [
                    "PO Number",
                    "Vendor Name",
                    "Order Date",
                    "Grand Total",
                    "Status",
                ]
                ws_po.append(po_headers)
                for po in all_pos.values():
                    ws_po.append(
                        [
                            po.po_number,
                            po.vendor_name,
                            po.order_date,
                            po.grand_total,
                            getattr(po, "status", "unknown"),
                        ]
                    )
            if all_grns:
                ws_grn = wb.create_sheet("Goods Receipt Notes")
                grn_headers = ["GRN Number", "PO Number", "Received Date", "Status"]
                ws_grn.append(grn_headers)
                for grn in all_grns.values():
                    ws_grn.append(
                        [
                            grn.grn_number,
                            grn.po_number,
                            grn.received_date,
                            getattr(grn, "status", "unknown"),
                        ]
                    )
            virtual_workbook = save_virtual_workbook(wb)
            return StreamingResponse(
                io.BytesIO(virtual_workbook),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f"attachment; filename=ap_export_{timestamp}.xlsx"
                },
            )
        except ImportError as e:
            raise HTTPException(
                status_code=501,
                detail=f"XLSX export requires openpyxl library. Please install it with 'pip install openpyxl'. Error: {str(e)}",
            )
        except Exception as e:
            raise HTTPException(
                status_code=500, detail=f"Failed to generate XLSX export: {str(e)}"
            )
    raise HTTPException(
        status_code=400,
        detail=f"Invalid export format '{export_format}'. Supported formats: csv, xlsx",
    )


# --- START MODIFICATION: Robust PO Update Logic ---
@router.put("/purchase-orders/{po_db_id}")
def update_purchase_order(
    po_db_id: int,
    request: schemas.PurchaseOrderUpdateRequest,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    try:
        po = (
            db.query(models.PurchaseOrder)
            .options(joinedload(models.PurchaseOrder.invoices))
            .filter(models.PurchaseOrder.id == po_db_id)
            .first()
        )
        if not po:
            raise HTTPException(status_code=404, detail="Purchase Order not found")
        po.version = request.version
        changes = request.changes

        # --- NEW MERGE LOGIC FOR LINE ITEMS ---
        if "line_items" in changes and isinstance(changes["line_items"], list):
            # Create a map of the new line item changes by description
            changes_map = {
                item.get("description"): item
                for item in changes["line_items"]
                if isinstance(item, dict)
            }

            # Get the current line items from the PO
            current_line_items = (
                po.line_items if isinstance(po.line_items, list) else []
            )

            # Create the new list by merging changes
            updated_line_items = []
            for existing_item in current_line_items:
                if (
                    isinstance(existing_item, dict)
                    and existing_item.get("description") in changes_map
                ):
                    # If this item was edited, merge the changes and add it
                    updated_item = {
                        **existing_item,
                        **changes_map[existing_item.get("description")],
                    }
                    updated_line_items.append(updated_item)
                else:
                    # Otherwise, keep the existing item as is
                    updated_line_items.append(existing_item)

            # Replace the line_items in the changes dictionary with the fully merged list
            changes["line_items"] = updated_line_items
        # --- END NEW MERGE LOGIC ---

        # Update the PO object with all changes
        for key, value in changes.items():
            if hasattr(po, key):
                setattr(po, key, value)

        if po.raw_data_payload:
            for key, value in changes.items():
                if key in po.raw_data_payload:
                    po.raw_data_payload[key] = value

        invoices_to_rematch = po.invoices
        summary_parts = []
        if "line_items" in changes:
            summary_parts.append(f"Updated {len(changes['line_items'])} line item(s).")
        other_changes = {k: v for k, v in changes.items() if k != "line_items"}
        if other_changes:
            summary_parts.append(f"Updated fields: {', '.join(other_changes.keys())}.")
        update_summary = (
            " ".join(summary_parts)
            if summary_parts
            else "No specific changes detailed."
        )

        for inv in invoices_to_rematch:
            log_audit_event(
                db=db,
                invoice_db_id=inv.id,
                user=current_user.email,
                action="PO Updated, Triggering Rematch",
                entity_type="Invoice",
                entity_id=inv.invoice_id,
                summary=f"PO {po.po_number} was updated. {update_summary}",
                details={"po_number": po.po_number, "changes": changes},
            )

        db.commit()
        for inv in invoices_to_rematch:
            background_tasks.add_task(matching_engine.run_match_for_invoice, db, inv.id)

        db.refresh(po)
        return {
            "message": "Purchase Order updated. Rematching related invoices in the background."
        }

    except StaleDataError:
        db.rollback()
        raise HTTPException(
            status_code=409,
            detail="This Purchase Order has been modified by someone else. Please refresh and try again.",
        )
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


# --- END MODIFICATION ---


@router.get("/jobs/{job_id}/invoices", response_model=List[schemas.Invoice])
def get_invoices_for_job(job_id: int, db: Session = Depends(get_db)):
    job = db.query(models.Job).filter(models.Job.id == job_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job ID not found")
    invoices = db.query(models.Invoice).filter(models.Invoice.job_id == job_id).all()
    return invoices if invoices else []


@router.post(
    "/create-po-from-invoice/{invoice_db_id}",
    response_model=schemas.PurchaseOrder,
    status_code=202,
)
def create_po_from_invoice(
    invoice_db_id: int,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    try:
        new_po = document_service.create_po_from_invoice_logic(
            db=db,
            invoice_db_id=invoice_db_id,
            current_user=current_user,
            background_tasks=background_tasks,
        )
        return new_po
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"An unexpected error occurred: {str(e)}"
        )


@router.post("/upload-policy", response_model=schemas.Job, status_code=202)
async def upload_policy_document(
    background_tasks: BackgroundTasks,
    files: List[UploadFile] = File(...),
    rule_level: str = QueryParam("system", description="Rule level: system, segment, or customer"),
    segment: Optional[str] = QueryParam(None, description="Segment name if rule_level is 'segment'"),
    customer_id: Optional[str] = QueryParam(None, description="Customer ID if rule_level is 'customer'"),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    """
    Upload policy documents and create rules at specified level (system, segment, or customer).
    """
    if not files:
        raise HTTPException(status_code=400, detail="No files were uploaded.")
    
    # Validate rule level and required parameters
    if rule_level not in ["system", "segment", "customer"]:
        raise HTTPException(status_code=400, detail="Rule level must be 'system', 'segment', or 'customer'")
    
    if rule_level == "segment" and not segment:
        raise HTTPException(status_code=400, detail="Segment name is required when rule_level is 'segment'")
    
    if rule_level == "customer" and not customer_id:
        raise HTTPException(status_code=400, detail="Customer ID is required when rule_level is 'customer'")
    
    # Ensure only PDF files are uploaded
    for file in files:
        if not file.filename.lower().endswith('.pdf'):
            raise HTTPException(status_code=400, detail=f"Only PDF files are allowed. {file.filename} is not a PDF.")
    
    try:
        os.makedirs(PDF_STORAGE_PATH, exist_ok=True)
    except OSError as e:
        raise HTTPException(
            status_code=500,
            detail=f"Cannot create upload directory: {e}. Check volume mounts and permissions.",
        )
    
    file_data_list: List[Dict[str, Any]] = []
    for file in files:
        try:
            content = await file.read()
            safe_filename = os.path.basename(file.filename)
            file_path = os.path.join(PDF_STORAGE_PATH, safe_filename)
            with open(file_path, "wb") as buffer:
                buffer.write(content)
            
            file_data_list.append({
                "filename": safe_filename,
                "content": content,
                "file_type": "loan_policy",
                "rule_level": rule_level,
                "segment": segment,
                "customer_id": customer_id,
                "source_folder": "policy_upload"
            })
        except (OSError, IOError) as e:
            raise HTTPException(
                status_code=500,
                detail=f"Failed to save file {file.filename}: {e}. Check disk space and permissions.",
            )
    
    job = models.Job(total_files=len(files))
    db.add(job)
    db.commit()
    db.refresh(job)
    
    # Use enhanced processing function that handles policy documents
    background_tasks.add_task(
        tasks_service.process_policy_documents, job.id, file_data_list
    )
    return job


@router.get("/failed-ingestions", response_model=List[FailedIngestionResponse])
def get_failed_ingestions(
    db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)
):
    return (
        db.query(models.FailedIngestion)
        .order_by(models.FailedIngestion.created_at.desc())
        .all()
    )


@router.get("/failed-ingestions/{failed_id}/data", response_model=Dict[str, Any])
def get_failed_ingestion_data(
    failed_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    failure = db.query(models.FailedIngestion).filter_by(id=failed_id).first()
    if not failure:
        raise HTTPException(
            status_code=404, detail="Failed ingestion record not found."
        )
    return failure.raw_data


@router.post("/failed-ingestions/{failed_id}/retry", status_code=200)
def retry_failed_ingestion(
    failed_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    failure = db.query(models.FailedIngestion).filter_by(id=failed_id).first()
    if not failure:
        raise HTTPException(
            status_code=404, detail="Failed ingestion record not found."
        )
    saved_obj, error = None, None
    if failure.document_type == "PurchaseOrder":
        saved_obj, error = ingestion_service._save_po_from_dict(
            db, failure.raw_data, failure.job_id
        )
    elif failure.document_type == "GoodsReceiptNote":
        saved_obj, error = ingestion_service._save_grn_from_dict(
            db, failure.raw_data, failure.job_id
        )
    else:
        raise HTTPException(
            status_code=400,
            detail=f"Cannot retry document type: {failure.document_type}",
        )
    if error:
        failure.error_message = error
        db.commit()
        raise HTTPException(status_code=400, detail=f"Retry failed: {error}")
    db.delete(failure)
    db.commit()
    return {
        "message": f"{failure.document_type} from {failure.filename} successfully ingested."
    }
